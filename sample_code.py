
import random                   # We cover random numbers in the
from time import sleep
import turtle
import openpyxl as xl
from utils import find_max
import math
import random

# Day 11

white_paper = """
Bitcoin: A Peer-to-Peer Electronic Cash SystemSatoshi Nakamotosatoshin@gmx.comwww.bitcoin.orgAbstract.   A  purely   peer-to-peer   version   of   electronic   cash   would   allow   onlinepayments   to   be   sent   directly   from   one   party   to   another   without   going   through   afinancial institution.   Digital signatures provide part of the solution, but the mainbenefits are lost if a trusted third party is still required to prevent double-spending.We propose a solution to the double-spending problem using a peer-to-peer network.The   network   timestamps   transactions   by   hashing   them   into   an   ongoing   chain   ofhash-based proof-of-work, forming a record that cannot be changed without redoingthe proof-of-work.   The longest chain not only serves as proof of the sequence ofevents witnessed, but proof that it came from the largest pool of CPU power.   Aslong as a majority of CPU power is controlled by nodes that are not cooperating toattack the network,  they'll  generate the  longest  chain  and  outpace attackers.   Thenetwork itself requires minimal structure.   Messages are broadcast on a best effortbasis,   and   nodes   can   leave   and   rejoin   the   network   at   will,   accepting   the   longestproof-of-work chain as proof of what happened while they were gone.1.IntroductionCommerce on the Internet has come to rely almost exclusively on financial institutions serving astrusted third  parties  to process electronic payments.   While the  system works  well enough formost   transactions,   it   still   suffers   from   the   inherent   weaknesses   of   the   trust   based   model.Completely non-reversible transactions are not really possible, since financial institutions cannotavoid   mediating   disputes.     The   cost   of   mediation   increases   transaction   costs,   limiting   theminimum practical transaction size and cutting off the possibility for small casual transactions,and   there   is   a   broader   cost   in   the   loss   of   ability   to   make   non-reversible   payments   for   non-reversible services.  With the possibility of reversal, the need for trust spreads.  Merchants mustbe wary of their customers, hassling them for more information than they would otherwise need.A certain percentage of fraud is accepted as unavoidable.  These costs and payment uncertaintiescan be avoided in person by using physical currency, but no mechanism exists to make paymentsover a communications channel without a trusted party.What is needed is an electronic payment system based on cryptographic proof instead of trust,allowing any two willing parties to transact directly with each other without the need for a trustedthird  party.    Transactions  that  are  computationally  impractical  to   reverse   would  protect  sellersfrom fraud, and routine escrow mechanisms could easily be implemented to protect buyers.   Inthis paper, we propose a solution to the double-spending problem using a peer-to-peer distributedtimestamp server to generate computational proof of the chronological order of transactions.  Thesystem   is   secure   as   long   as   honest   nodes   collectively   control   more   CPU   power   than   anycooperating group of attacker nodes.1
2.TransactionsWe define an electronic coin as a chain of digital signatures.  Each owner transfers the coin to thenext by digitally signing a hash of the previous transaction and the public key of the next ownerand adding these to the end of the coin.  A payee can verify the signatures to verify the chain ofownership.The problem of course is the payee can't verify that one of the owners did not double-spendthe coin.  A common solution is to introduce a trusted central authority, or mint, that checks everytransaction for double spending.  After each transaction, the coin must be returned to the mint toissue a new coin, and only coins issued directly from the mint are trusted not to be double-spent.The   problem   with   this   solution   is   that   the   fate   of   the   entire   money   system   depends   on   thecompany running the mint, with every transaction having to go through them, just like a bank.We   need   a   way   for   the   payee   to   know   that   the   previous   owners   did   not   sign   any   earliertransactions.   For our purposes, the earliest transaction is the one that counts, so we don't careabout later attempts to double-spend.  The only way to confirm the absence of a transaction is tobe aware of all transactions.  In the mint based model, the mint was aware of all transactions anddecided   which   arrived   first.    To  accomplish   this   without   a   trusted   party,   transactions   must   bepublicly announced [1], and we need a system for participants to agree on a single history of theorder in which they were received.  The payee needs proof that at the time of each transaction, themajority of nodes agreed it was the first received. 3.Timestamp ServerThe solution we propose begins with a timestamp server.  A timestamp server works by taking ahash   of   a   block   of   items   to   be   timestamped   and   widely   publishing   the   hash,   such   as   in   anewspaper or Usenet post [2-5].   The timestamp proves that the data must have existed at thetime, obviously, in order to get into the hash.  Each timestamp includes the previous timestamp inits hash, forming a chain, with each additional timestamp reinforcing the ones before it.2BlockItemItem...HashBlockItemItem...HashTransactionOwner 1'sPublic KeyOwner 0'sSignatureHashTransactionOwner 2'sPublic KeyOwner 1'sSignatureHashVerifyTransactionOwner 3'sPublic KeyOwner 2'sSignatureHashVerifyOwner 2'sPrivate KeyOwner 1'sPrivate KeySign  Sign  Owner 3'sPrivate Key
4.Proof-of-WorkTo implement a distributed timestamp server on a peer-to-peer basis, we will need to use a proof-of-work system  similar to Adam  Back's  Hashcash  [6],  rather than  newspaper  or  Usenet  posts.The proof-of-work involves scanning for a value that when hashed, such as with SHA-256, thehash begins with a number of zero bits.  The average work required is exponential in the numberof zero bits required and can be verified by executing a single hash.For our timestamp network, we implement the proof-of-work by incrementing a nonce in theblock until a value is found that gives the block's hash the required zero bits.   Once the CPUeffort   has   been   expended   to   make   it   satisfy   the   proof-of-work,   the   block   cannot   be   changedwithout  redoing  the   work.    As   later   blocks   are  chained   after  it,   the  work  to  change  the  blockwould include redoing all the blocks after it.The proof-of-work also solves the problem of determining representation in majority decisionmaking.  If the majority were based on one-IP-address-one-vote, it could be subverted by anyoneable   to   allocate   many   IPs.     Proof-of-work   is   essentially   one-CPU-one-vote.     The   majoritydecision is represented by the longest chain, which has the greatest proof-of-work effort investedin it.  If a majority of CPU power is controlled by honest nodes, the honest chain will grow thefastest and outpace any competing chains.   To modify a past block, an attacker would have toredo the proof-of-work of the block and all blocks after it and then catch up with and surpass thework of the honest nodes.  We will show later that the probability of a slower attacker catching updiminishes exponentially as subsequent blocks are added.To compensate for increasing hardware speed and varying interest in running nodes over time,the proof-of-work difficulty is determined by a moving average targeting an average number ofblocks per hour.  If they're generated too fast, the difficulty increases.5.NetworkThe steps to run the network are as follows:1)New transactions are broadcast to all nodes.2)Each node collects new transactions into a block.  3)Each node works on finding a difficult proof-of-work for its block.4)When a node finds a proof-of-work, it broadcasts the block to all nodes.5)Nodes accept the block only if all transactions in it are valid and not already spent.6)Nodes express their acceptance of the block by working on creating the next block in thechain, using the hash of the accepted block as the previous hash.Nodes   always   consider   the   longest   chain   to   be   the   correct   one   and   will   keep   working   onextending it.   If two nodes broadcast different versions of the next block simultaneously, somenodes may receive one or the other first.  In that case, they work on the first one they received,but save the other branch in case it becomes longer.  The tie will be broken when the next proof-of-work   is   found   and   one   branch   becomes   longer;   the   nodes   that   were   working   on   the   otherbranch will then switch to the longer one.3BlockPrev HashNonceTxTx...BlockPrev HashNonceTxTx...
New transaction broadcasts do not necessarily need to reach all nodes.  As long as they reachmany nodes, they will get into a block before long.  Block broadcasts are also tolerant of droppedmessages.  If a node does not receive a block, it will request it when it receives the next block andrealizes it missed one.6.IncentiveBy convention, the first transaction in a block is a special transaction that starts a new coin ownedby the creator of the block.  This adds an incentive for nodes to support the network, and providesa way to initially distribute coins into circulation, since there is no central authority to issue them.The steady addition of a constant of amount of new coins is analogous to gold miners expendingresources to add gold to circulation.  In our case, it is CPU time and electricity that is expended.The incentive can also be funded with transaction fees.  If the output value of a transaction isless than its input value, the difference is a transaction fee that is added to the incentive value ofthe   block   containing   the   transaction.     Once   a   predetermined   number   of   coins   have   enteredcirculation, the incentive can transition entirely to transaction fees and be completely inflationfree.The   incentive   may   help   encourage   nodes   to   stay   honest.     If   a   greedy   attacker   is   able   toassemble more CPU power than all the honest nodes, he would have to choose between using itto defraud people by stealing back his payments, or using it to generate new coins.  He ought tofind it more profitable to play by the rules, such rules that favour him with more new coins thaneveryone else combined, than to undermine the system and the validity of his own wealth.7.Reclaiming Disk SpaceOnce the latest transaction in a coin is buried under enough blocks, the spent transactions beforeit   can   be   discarded   to   save   disk   space.     To   facilitate   this   without   breaking   the   block's   hash,transactions are hashed in a Merkle Tree [7][2][5], with only the root included in the block's hash.Old blocks can then be compacted by stubbing off branches of the tree.   The interior hashes donot need to be stored.A  block   header   with   no   transactions   would   be   about   80   bytes.     If   we   suppose   blocks   aregenerated every 10 minutes, 80 bytes * 6 * 24 * 365 = 4.2MB per year.  With computer systemstypically selling with 2GB of RAM as of 2008, and Moore's Law predicting current growth of1.2GB   per   year,   storage   should   not   be   a   problem   even   if   the   block   headers   must   be   kept   inmemory.4BlockBlockBlock Header (Block Hash)Prev HashNonceHash01Hash0Hash1Hash2Hash3Hash23Root HashHash01Hash2Tx3Hash23Block Header (Block Hash)Root HashTransactions Hashed in a Merkle TreeAfter Pruning Tx0-2 from the BlockPrev HashNonceHash3Tx0Tx1Tx2Tx3
8.Simplified Payment VerificationIt is possible to verify payments without running a full network node.  A user only needs to keepa copy of the block headers of the longest proof-of-work chain, which he can get by queryingnetwork   nodes   until   he's   convinced   he   has   the   longest   chain,   and   obtain   the   Merkle   branchlinking   the   transaction   to   the   block   it's   timestamped   in.     He   can't   check   the   transaction   forhimself, but by linking it to a place in the chain, he can see that a network node has accepted it,and blocks added after it further confirm the network has accepted it.As such, the verification is reliable as long as honest nodes control the network, but is morevulnerable   if   the   network   is   overpowered   by   an   attacker.     While   network   nodes   can   verifytransactions   for   themselves,   the   simplified   method   can   be   fooled   by   an   attacker's   fabricatedtransactions for as long as the attacker can continue to overpower the network.   One strategy toprotect against this would be to accept alerts from network nodes when they detect an invalidblock,   prompting   the   user's   software   to   download   the   full   block   and   alerted   transactions   toconfirm the inconsistency.  Businesses that receive frequent payments will probably still want torun their own nodes for more independent security and quicker verification.9.Combining and Splitting ValueAlthough   it   would   be   possible   to   handle   coins   individually,   it   would   be   unwieldy   to   make   aseparate   transaction   for   every   cent   in   a   transfer.     To   allow   value   to   be   split   and   combined,transactions  contain  multiple  inputs  and  outputs.    Normally  there will  be either  a  single  inputfrom a larger previous transaction or multiple inputs combining smaller amounts, and at most twooutputs: one for the payment, and one returning the change, if any, back to the sender.  It should be noted that fan-out, where a transaction depends on several transactions, and thosetransactions depend on many more, is not a problem here.   There is never the need to extract acomplete standalone copy of a transaction's history.5TransactionIn...InOut...Hash01Hash2Hash3Hash23Block HeaderMerkle RootPrev HashNonceBlock HeaderMerkle RootPrev HashNonceBlock HeaderMerkle RootPrev HashNonceMerkle Branch for Tx3Longest Proof-of-Work ChainTx3
10.PrivacyThe traditional banking model achieves a level of privacy by limiting access to information to theparties involved and the trusted third party.   The necessity to announce all transactions publiclyprecludes this method, but privacy can still be maintained by breaking the flow of information inanother place: by keeping public keys anonymous.   The public can see that someone is sendingan amount to someone else, but without information linking the transaction to anyone.   This issimilar   to   the   level   of   information   released   by   stock   exchanges,   where   the   time   and   size   ofindividual trades, the "tape", is made public, but without telling who the parties were.As an additional firewall, a new key pair should be used for each transaction to keep themfrom   being   linked   to   a   common   owner.     Some   linking   is   still   unavoidable   with   multi-inputtransactions, which necessarily reveal that their inputs were owned by the same owner.  The riskis that if the owner of a key is revealed, linking could reveal other transactions that belonged tothe same owner.11.CalculationsWe consider the scenario of an attacker trying to generate an alternate chain faster than the honestchain.  Even if this is accomplished, it does not throw the system open to arbitrary changes, suchas creating value out of thin air or taking money that never belonged to the attacker.  Nodes arenot going to accept an invalid transaction as payment, and honest nodes will never accept a blockcontaining them.   An attacker can only try to change one of his own transactions to take backmoney he recently spent.The race between the honest chain and an attacker chain can be characterized as a BinomialRandom Walk.  The success event is the honest chain being extended by one block, increasing itslead by +1, and the failure event is the attacker's chain being extended by one block, reducing thegap by -1.The probability of an attacker catching up from a given deficit is analogous to a Gambler'sRuin problem.  Suppose a gambler with unlimited credit starts at a deficit and plays potentially aninfinite   number   of   trials   to   try   to   reach   breakeven.    We   can   calculate   the   probability   he   everreaches breakeven, or that an attacker ever catches up with the honest chain, as follows [8]:p = probability an honest node finds the next blockq = probability the attacker finds the next blockqz = probability the attacker will ever catch up from z blocks behindqz={1ifp≤qq/pzifpq}6IdentitiesTransactionsTrustedThird PartyCounterpartyPublicIdentitiesTransactionsPublicNew Privacy ModelTraditional Privacy Model
# include <math.h>double AttackerSuccessProbability(double q, int z){    double p = 1.0 - q;    double lambda = z * (q / p);    double sum = 1.0;    int i, k;    for (k = 0; k <= z; k++)    {        double poisson = exp(-lambda);        for (i = 1; i <= k; i++)            poisson *= lambda / i;        sum -= poisson * (1 - pow(q / p, z - k));    }    return sum;}7
Given our assumption that p > q, the probability drops exponentially as the number of blocks theattacker has to catch up with increases.   With the odds against him, if he doesn't make a luckylunge forward early on, his chances become vanishingly small as he falls further behind.We   now   consider   how   long   the   recipient   of  a   new   transaction   needs   to   wait   before   beingsufficiently certain the sender can't change the transaction.  We assume the sender is an attackerwho wants to make the recipient believe he paid him for a while, then switch it to pay back tohimself  after   some   time   has   passed.    The   receiver   will   be   alerted   when   that   happens,   but   thesender hopes it will be too late.The receiver generates a new key pair and gives the public key to the sender shortly beforesigning.  This prevents the sender from preparing a chain of blocks ahead of time by working onit continuously until he is lucky enough to get far enough ahead, then executing the transaction atthat  moment.   Once  the transaction is   sent,  the dishonest  sender starts  working  in  secret on  aparallel chain containing an alternate version of his transaction.The recipient waits until the transaction has been added to a block and  z  blocks have beenlinked   after   it.     He   doesn't   know   the   exact   amount   of   progress   the   attacker   has   made,   butassuming   the   honest   blocks   took   the   average   expected   time   per   block,   the   attacker's   potentialprogress will be a Poisson distribution with expected value:=zqpTo get the probability the attacker could still catch up now, we multiply the Poisson density foreach amount of progress he could have made by the probability he could catch up from that point:∑k=0∞ke−k!⋅{q/pz−kifk≤z1ifkz}Rearranging to avoid summing the infinite tail of the distribution...1−∑k=0zke−k!1−q/pz−kConverting to C code...
Running some results, we can see the probability drop off exponentially with z.q=0.1z=0    P=1.0000000z=1    P=0.2045873z=2    P=0.0509779z=3    P=0.0131722z=4    P=0.0034552z=5    P=0.0009137z=6    P=0.0002428z=7    P=0.0000647z=8    P=0.0000173z=9    P=0.0000046z=10   P=0.0000012q=0.3z=0    P=1.0000000z=5    P=0.1773523z=10   P=0.0416605z=15   P=0.0101008z=20   P=0.0024804z=25   P=0.0006132z=30   P=0.0001522z=35   P=0.0000379z=40   P=0.0000095z=45   P=0.0000024z=50   P=0.0000006Solving for P less than 0.1%...P < 0.001q=0.10   z=5q=0.15   z=8q=0.20   z=11q=0.25   z=15q=0.30   z=24q=0.35   z=41q=0.40   z=89q=0.45   z=34012.ConclusionWe have proposed a system for electronic transactions without relying on trust.  We started withthe   usual   framework   of   coins   made   from   digital   signatures,   which   provides   strong   control   ofownership,   but   is   incomplete   without   a   way   to   prevent   double-spending.     To   solve   this,   weproposed a peer-to-peer network using proof-of-work to record a public history of transactionsthat   quickly   becomes   computationally   impractical   for   an   attacker   to   change   if   honest   nodescontrol a majority of CPU power.   The network is robust in its unstructured simplicity.   Nodeswork all at once with little coordination.   They do not need to be identified, since messages arenot routed to any particular place and only need to be delivered on a best effort basis.  Nodes canleave   and   rejoin   the   network   at   will,   accepting   the   proof-of-work   chain   as   proof   of   whathappened while they were gone.  They vote with their CPU power, expressing their acceptance ofvalid blocks by working on extending them and rejecting invalid blocks by refusing to work onthem.  Any needed rules and incentives can be enforced with this consensus mechanism.
"""
white_paper_words = white_paper.split()


def word_count(word):
    counter = 0
    for i in white_paper_words:
        if i == word:
            counter += 1
    return counter


print(word_count("crypto"))  # zero times
print(word_count("cash"))   # one time
print(word_count("price"))  # zero times


###

words = ("apple", "banana", "trees", "car", "optopus", "bitcoin", "hashtable")


def aleast_5_letters(words):
    counter = 0
    for word in words:
        if len(word) >= 5:
            counter += 1
    return counter


print(aleast_5_letters(words))

###

nums = (1, 3, 4, 5, 656, 25, 675, 123, 78)

# sums up Only negitive numbers


def add_neg_numbs(nums):
    sum = 0
    for i in nums:
        if i % 2 == 1:
            sum = sum + i
    return sum


# adds up all the numbers
def add_numbs(nums):
    sum = 0
    for i in nums:
        sum = sum + i
    return sum


print(add_numbs(nums))

# counts the number of neg numbers


def if_neg(nums):
    counter = 0
    for i in nums:
        if i % 2 == 1:
            counter += 1
    return counter


print(if_neg(nums))


###
# did NOT write code below

rng = random.Random()  # modules chapter, so peek ahead.
number = rng.randrange(1, 1000)  # Get random number between [1 and 1000).

guesses = 0
msg = ""

while True:
    guess = int(input(msg + "\nGuess my number between 1 and 1000: "))
    guesses += 1
    if guess > number:
        msg += str(guess) + " is too high.\n"
    elif guess < number:
        msg += str(guess) + " is too low.\n"
    else:
        break

input("\n\nGreat, you got it in {0} guesses!\n\n".format(guesses))


# Day 10


xs = [12, 10, 32, 3, 66, 17, 42, 99, 20]

# for a in xs:
#   print(a)

# for a in xs:
#    print(f"{a}, {a ** 2}")

# total = 0
# for a in xs:
#    total += a

# print(total)

product = 1
for a in xs:
    product *= a

print(product)


###
months = ["Jan", "Feb", "March", "April", "May",
          "June", "July", "Aug", "Sept", "Oct", "Nov", "Dec"]

for m in months:
    print(f"{m} is a month of the year.")

###
print("We like Python's turtles!" * 1000)

# OR as DEF, below

string = "hello world  "
x = 40


def print_string(string, x):
    press = ""
    press = string * x
    return press


print(print_string(string, x))


# Day 9

###


a = 2
while True:
    print(a)
    sleep(.5)
    a **= 2

###


want_bg_color = input("what color background do u want: ")
wn = turtle.Screen()
wn.bgcolor(want_bg_color)      # Set the window background color
wn.title("Hello, Tess!")      # Set the window title

want_turtle_color = input("what color turtle do u want: ")
tess = turtle.Turtle()
tess.color(want_turtle_color)
# Tell tess to change her color

want_size_pen = int(input("what size pen do u want: "))
tess.pensize(want_size_pen)               # Tell tess to set her pen width

tess.forward(50)
tess.left(120)
tess.forward(50)

wn.mainloop()


def hour_of_alarm(now, hours_waiting):
    alarm_hour = 0
    alarm_hour = (now + hours_waiting) % 24
    return alarm_hour


print(hour_of_alarm(3, 49))


# WORKING !

now = 14
alarm = (now + 51) % 24
print(alarm)

#

word = "apple"
new_word = ""


def duplicate_encode(word):
    new_word = ""

    for index in word:
        if index in word:
            index = ")"
            new_word += index
        else:
            index = "("
            new_word += index
    return(new_word)


print(new_word)


# Day 8


wb = xl.load_workbook("store.xlsx")

sheet = wb["Sheet1"]

cell = sheet["a1"]
cell = sheet.cell(1, 1)

print(sheet.max_row)

# Day 7


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())


dice = ["1", "2", "3", "4", "5", "6"]

lucky1 = random.choice(dice)

lucky2 = random.choice(dice)

print(f"({lucky1}, {lucky2})")


for i in range(2):
    print((random.randint(1, 6))

# for i in range(3):
#   print(random.randint(10, 30))

peps=["alice", "bob", "cathy", "jay"]

leader=random.choice(peps)

print(leader)


def find_max(numbers):

    max=numbers[0]
    for number in numbers:
        if number > max:
            max=number
    return(max)


numbers=[11, 5, 7, 15, 3]

max=find_max(numbers)

print(max)


#  import converters
# all

#   from converters import kg_to_lbs
# just some


class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print("bark")


class Cat(Mammal):
    pass


dog1=Dog()
dog1.walk()


class Person:
    def __init__(self, name):
        self.name=name

    def talk(self):
        print(f"hi, i'm {self.name}.")


john=Person("john Smith")
john.talk()


try:
    age=int(input("age: "))
    print(age)
except ValueError:
    print("that's not a valid age")


# Day 6

# create functions


def emoji_converter(message):
    words=message.split(" ")

    emoji={
        ": )": "😃",
        ": (": "😞"
    }

    output=""
    for word in words:
        output += emoji.get(word, word) + " "
    return output


message=input(">")
print(emoji_converter(message))


# dictionaries

message=input(">")
words=message.split(" ")

emoji={
    ": )": "test",
    ": (": "😞"
}

output=""
for word in words:
    output += emoji.get(word, word) + " "

print(output)
# BROKEN !!!!!


numbers={
    "1": "one",
    "2": "two",
    "3": "three",
    "4": "four",
    "5": "five"
}

output=""
digi=input("what's your number: ")

for num in digi:
    output += numbers.get(num, "!") + " "
print(output)


clients={
    "name": "Joe",
    "age": 34,
    "is_hodl": True,
}
print(clients["name"])
print(clients.get("age"))


#  tumple = (1, 2, 3)
#  x, y, z = tumple
#  print(x)


numbers=[1, 444, 7, 7, 67, 345, 7]
unques=[]

for num in numbers:
    if num not in unques:
        unques.append(num)
print(unques)


num_list=[1, 444, 67, 345, 7]

max=num_list[0]

for num in num_list:
    if num > max:
        max=num
print(max)

# Day 5

numbers=[5, 2, 5, 2, 2]

for x_count in numbers:
    output=''
    for count in range(x_count):
        output += 'x'
    print(output)


prices=[10, 20, 30]
total=0

for cost in prices:
    total += cost
print(f"the total is {total})


user_says=""
now_started=False

while True:
    user_says=input(">").lower()

    if user_says == "start":
        if now_started:
            print("your already started!!!")
        else:
            now_started=True
            print("your car is starting...")

    elif user_says == "stop":
        if not now_started:
            print("your already stopped!!!")
        else:
            now_started=False
            print("your car is stoping...")

    elif user_says == "help":
        print("""
        start - starts car
        stops - stops car
        quit - quits game
        """)
    elif user_says == "quit":
        break

    else:
        print("sorry, i don't understand")


magic_number=7
guess_limit=3

num_of_trys=0


while num_of_trys < guess_limit:
    user_guess=int(input("please guess 1-10: "))
    num_of_trys += 1

    if user_guess == magic_number:
        print("your right!")
        break
    else:
        print("you lose")

print("game over")


# Day 4

weight=int(input("weight: "))
unit=input("(L)bs or (K)g: ")

if unit.upper() == "L":
    converted=weight * .45
    print(f"you r {converted} kgs.")
else:
    converted=weight / .45
    print(f"you r {converted} pounds.")


name="jay"

if len(name) < 3:
    print("your names too short")
elif len(name) >= 50:
    print("yo names too long, dude")
else:
    print(f"thnaks {name}!")


house_price=1000000
has_good_credit=True

if has_good_credit:
    deposit=house_price * .1

else:
    deposit=house_price * .2

print(f"the despoit is ${deposit}.")


lbs=input("how many pound u weigh? ")
kilos=int(lbs) / 2
KGs=kilos * .1
KG=kilos - KGs

print("you weigh " + str(KG) + " kg...wow!")


name=input("what's your name? ")
age=input("what's your age? ")
print(name + " is " + age + " years old.")
print("ok")
print("good")


def fizz_buzz(input):

    if (input % 5 == 0) and (input % 3 == 0):
        return("FizzBuzz")
    if input % 3 == 0:
        return("Fizz")
    if input % 5 == 0:
        return("Buzz")
    return(input)


print(fizz_buzz(34))


def multing_unknown_numbers(*abc):

    total=1
    for x in abc:
        total *= x   # total = total * x
    return total


print(multing_unknown_numbers(3, 9,))


def greet(first_name, last_name):
    print(f"Hi {first_name} {last_name}")
    print("welcome aboard")


greet("Jay", "Oceans")


# Day 3

# add change
# did the file name change ?...
# test if git is needed ... NO

count=0
for number in range(1, 30):
    if number % 2 == 0:
        print(number)
        count += 1
print(f"we have {count} numbers...")


# Day 2 code disappeared (sad)

# LOST CODE...


# Day 1

# x = input("x: ")
y=int(x) + 3
print(y)

temp=25
if temp < 20:
    print("it's cold!")
elif temp > 30:
    print("it's crazy hot!!")
else:
    print("it's fine")
print("ok")

age=24
if age < 18:
    print("too young!")
else:
    print("ok!")

# OR
age=12

message=("ok") if age > 18 else "too young"

print(message)

# and  not  or

stack_sats=True
strong_hands=True
big_ego=False

if (stack_sats or strong_hands) and not big_ego:
    print("to the moon!")
else:
    print("sad face...")


# coding bitcoin

shit_coin=False
strong_hands=False
big_heart=False

if (strong_hands or big_heart) and not shit_coin:
    print("to the moon!")
else:
    print("game over")

    # chaining comparison operators

temp=76

if 50 < temp <= 75:
    print("it's a nice day")
else:
    print("it's too hot or too cold")

# for  in a range

for number in range(1, 100, 2):
    print("Try", number, number * "*")

# maybe trying to get user info OR send something

successful_sent=True

for number in range(1, 4):
    print("Try", number)
    if successful_sent:
        print("thanks")
        break
else:
    print("sorry, no go")

# Day 0

print("Hello World !")
print("*" * 10)
print("hello")

x="Hello \n Sally"
big=5

print(big)

print """
hi,
this is Bob
how r u...
"""

print(len(x))
print(x[0])
print(x[1:4])
print(x[-4])

first="Jay"
last=" Oceans"
full=first + last
print(full)

bold_first_name=first.upper()

print(bold_first_name)

print(full.find("O"))

print(full.find("o"))

print(full.replace("O", "F"))

print("Jay" in full)            # True

print("Joe" not in full)        # True

print(math.ceil(2))
